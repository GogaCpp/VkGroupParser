import vk_api
import datetime
import argparse
import openpyxl

def main():
  
    token="токен"
    parser = argparse.ArgumentParser(description="Сбор информации с группы VK")
    parser.add_argument("group_url", help="Ссылка на группу VK")
    parser.add_argument("start_date", help="Начало периода (YYYY-MM-DD)")
    parser.add_argument("output_file", help="Путь до файла с результатами (.xlsx)")
    args = parser.parse_args()

    group_url = args.group_url
    start_date = datetime.datetime.strptime(args.start_date, "%Y-%m-%d").date()
    output_file = args.output_file
    
    
    vk_session = vk_api.VkApi(token=token)
    vk = vk_session.get_api()
    
    group_id=get_group_id(vk,group_url)
    posts=get_posts(vk,group_id,start_date)
    
    wb = openpyxl.Workbook()
    posts_sheet=wb.active
    posts_sheet.title="Посты"
    comments_sheet=wb.create_sheet("Коментарии")
    likes_sheet=wb.create_sheet("Лайки")
    
    posts_sheet.append(["post_id", "текст", "дата", "кол-во лайков", "кол-во комментов"])
    comments_sheet.append(["post_id", "user_id", "текст", "дата"])
    likes_sheet.append(["post_id", "user_id"])
    
    for post in posts:
        post_id=post["id"]
        post_text=post["text"]
        post_date=datetime.datetime.fromtimestamp(post["date"]).strftime("%Y-%m-%d %H:%M:%S")
        post_likes_count=post["likes"]["count"]
        post_comments_count=post["comments"]["count"]
        posts_sheet.append([post_id,post_text,post_date,post_likes_count,post_comments_count])
        
        comments_list=get_comments(vk,group_id,post_id)
        for comment in comments_list:
            comment_user_id = comment["from_id"]
            comment_text=comment["text"]
            comment_date=datetime.datetime.fromtimestamp(comment["date"]).strftime("%Y-%m-%d %H:%M:%S")
            comments_sheet.append([post_id,comment_user_id,comment_text,comment_date])
        
        likes_list=get_likes(vk,group_id, post_id)
        for like in likes_list:
            like_user_id = like
            likes_sheet.append([post_id, like_user_id])
            
    wb.save(output_file)
    
def get_group_id(vk,group_url):
    parts = group_url.split("/")
    group_id = parts[-1].strip()
    if group_id.startswith("club"):
        group_id = group_id[4:]
        
    response = vk.groups.getById(group_id=group_id, fields='id', v='5.134')
    return response[0]['id']

def get_posts(vk,group_id,start_date):
    posts = []
    count=100;
    offset=0;
    while True:
        
        response=vk.wall.get(
          owner_id=-int(group_id),
          offset=offset,
          count=count,
          filter="all"
          )
        
        for post in response["items"]:
            
            if datetime.datetime.fromtimestamp(post["date"]).date() >= start_date:
                posts.append(post)
            else:
                return posts
            if len(response["items"])<count:
                break
            offset+=count
    
    return posts
  
def get_comments(vk,group_id,post_id):
    comments = []
    count=100;
    offset=0;
    while True:
        
        response=vk.wall.getComments(
          owner_id=-int(group_id),
          post_id=post_id,
          offset=offset,
          count=count
          )
        comments.extend(response["items"])
        if len(response["items"]) < count:
            break
        offset+=count
    return comments
    
def get_likes(vk,group_id, post_id):
    likes = []
    count = 100
    offset = 0
    while True:
      
        response = vk.likes.getList(
            type="post",
            owner_id=-int(group_id),
            item_id=post_id,
            count=count,
            offset=offset
        )
        likes.extend(response["items"])
        offset += count
        if len(response["items"]) < count:
            break
    return likes

    
    
    

if __name__ == "__main__":
    main()
